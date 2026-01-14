"""
Excel export service for formatting and exporting event data.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Optional
from datetime import datetime
from io import BytesIO
from loguru import logger

from app.models import EventData, EventType


class ExcelExporter:
    """
    Service for exporting event data to formatted Excel files.
    Creates professional-looking Excel workbooks with proper styling.
    """
    
    # Color scheme
    HEADER_COLOR = "366092"  # Dark blue
    ALT_ROW_COLOR = "F2F2F2"  # Light gray
    LINK_COLOR = "0563C1"    # Blue for hyperlinks
    
    def __init__(self):
        """Initialize the Excel exporter."""
        logger.info("ExcelExporter initialized")
    
    def _create_header_style(self) -> dict:
        """
        Create header cell styling.
        
        Returns:
            Dictionary of style attributes
        """
        return {
            'font': Font(bold=True, color="FFFFFF", size=11),
            'fill': PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid"),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def _create_cell_style(self, is_alt_row: bool = False) -> dict:
        """
        Create data cell styling.
        
        Args:
            is_alt_row: Whether this is an alternating row (for zebra striping)
        
        Returns:
            Dictionary of style attributes
        """
        style = {
            'alignment': Alignment(vertical='top', wrap_text=True),
            'border': Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
        }
        
        if is_alt_row:
            style['fill'] = PatternFill(start_color=self.ALT_ROW_COLOR, end_color=self.ALT_ROW_COLOR, fill_type="solid")
        
        return style
    
    def _apply_style(self, cell, style_dict: dict):
        """
        Apply style dictionary to a cell.
        
        Args:
            cell: Excel cell object
            style_dict: Dictionary of style attributes
        """
        for attr, value in style_dict.items():
            setattr(cell, attr, value)
    
    def _auto_adjust_column_widths(self, worksheet, min_width: int = 10, max_width: int = 50):
        """
        Auto-adjust column widths based on content.
        
        Args:
            worksheet: Excel worksheet
            min_width: Minimum column width
            max_width: Maximum column width
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_list(self, items: List[str]) -> str:
        """
        Format a list of strings for Excel display.
        
        Args:
            items: List of strings
        
        Returns:
            Formatted string
        """
        if not items:
            return ""
        return ", ".join(items)
    
    def _format_date(self, date: datetime) -> str:
        """
        Format datetime for Excel display.
        
        Args:
            date: Datetime object
        
        Returns:
            Formatted date string
        """
        if not date:
            return ""
        return date.strftime("%Y-%m-%d %H:%M")
    
    def _sanitize_datetime_string(self, datetime_str: str) -> str:
        """
        Convert ISO datetime string to Excel-compatible format (remove timezone).
        
        Args:
            datetime_str: ISO format datetime string (may include timezone)
        
        Returns:
            Formatted datetime string without timezone info
        """
        if not datetime_str:
            return ""
        
        try:
            # Handle various datetime formats
            from datetime import datetime as dt
            
            # Try parsing as ISO format with timezone
            # Format examples:
            # - "2026-01-14T10:30:00+00:00"
            # - "2026-01-14T10:30:00Z"
            # - "2026-01-13 14:57:46+00:00"
            
            # Replace 'Z' with '+00:00' for consistent parsing
            datetime_str_clean = str(datetime_str).replace('Z', '+00:00')
            
            # Remove timezone info if present (everything after + or -)
            if '+' in datetime_str_clean:
                datetime_str_clean = datetime_str_clean.split('+')[0]
            elif datetime_str_clean.count('-') > 2:  # More than date separators
                # Find last occurrence of - (timezone indicator)
                parts = datetime_str_clean.rsplit('-', 1)
                if len(parts[1]) <= 5:  # Likely timezone like -05:00
                    datetime_str_clean = parts[0]
            
            # Replace T with space if present
            datetime_str_clean = datetime_str_clean.replace('T', ' ')
            
            # Parse the cleaned string
            parsed_dt = dt.fromisoformat(datetime_str_clean)
            
            # Format for Excel
            return parsed_dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception as e:
            logger.warning(f"Failed to parse datetime '{datetime_str}': {e}")
            # Return original string if parsing fails
            return str(datetime_str)
    
    def create_events_workbook(
        self,
        events: List[EventData],
        include_metadata: bool = True
    ) -> Workbook:
        """
        Create an Excel workbook with event data.
        
        Args:
            events: List of EventData objects
            include_metadata: Whether to include a metadata sheet
        
        Returns:
            Workbook object
        """
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create main events sheet
        self._create_events_sheet(wb, events)
        
        # Create summary sheet
        if include_metadata:
            self._create_summary_sheet(wb, events)
        
        logger.info(f"Created Excel workbook with {len(events)} events")
        return wb
    
    def _create_events_sheet(self, workbook: Workbook, events: List[EventData]):
        """
        Create the main events data sheet with all required columns.
        
        Args:
            workbook: Workbook object
            events: List of EventData objects
        """
        ws = workbook.create_sheet("Events", 0)
        
        # Define headers - ALL REQUIRED FIELDS in specified order
        headers = [
            "Event Title",
            "Summary",
            "Event Type",
            "Perpetrator",
            "Location (Full Text)",
            "City",
            "Region/State",
            "Country",
            "Event Date",
            "Event Time",
            "Individuals Involved",
            "Organizations Involved",
            "Casualties (Killed)",
            "Casualties (Injured)",
            "Source Name",
            "Source URL",
            "Article Publication Date",
            "Extraction Confidence"
        ]
        
        # Write headers with styling
        header_style = self._create_header_style()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, header_style)
        
        # Write data rows
        for row_idx, event in enumerate(events, 2):
            is_alt_row = (row_idx % 2) == 0
            cell_style = self._create_cell_style(is_alt_row)
            
            col = 1
            
            # 1. Event Title
            cell = ws.cell(row=row_idx, column=col, value=event.title)
            self._apply_style(cell, cell_style)
            cell.font = Font(bold=True)
            col += 1
            
            # 2. Summary
            cell = ws.cell(row=row_idx, column=col, value=event.summary)
            self._apply_style(cell, cell_style)
            col += 1
            
            # 3. Event Type
            cell = ws.cell(row=row_idx, column=col, value=event.event_type.value.upper().replace("_", " "))
            self._apply_style(cell, cell_style)
            col += 1
            
            # 4. Perpetrator
            cell = ws.cell(row=row_idx, column=col, value=event.perpetrator or "")
            self._apply_style(cell, cell_style)
            col += 1
            
            # 5. Location (Full Text)
            location_full = str(event.location) if event.location else ""
            cell = ws.cell(row=row_idx, column=col, value=location_full)
            self._apply_style(cell, cell_style)
            col += 1
            
            # 6. City
            city = event.location.city if event.location else ""
            cell = ws.cell(row=row_idx, column=col, value=city or "")
            self._apply_style(cell, cell_style)
            col += 1
            
            # 7. Region/State
            region = event.location.region if event.location else ""
            cell = ws.cell(row=row_idx, column=col, value=region or "")
            self._apply_style(cell, cell_style)
            col += 1
            
            # 8. Country
            country = event.location.country if event.location else ""
            cell = ws.cell(row=row_idx, column=col, value=country or "")
            self._apply_style(cell, cell_style)
            col += 1
            
            # 9. Event Date
            event_date_str = ""
            if event.event_date:
                event_date_str = event.event_date.strftime("%Y-%m-%d")
            cell = ws.cell(row=row_idx, column=col, value=event_date_str)
            self._apply_style(cell, cell_style)
            col += 1
            
            # 10. Event Time
            cell = ws.cell(row=row_idx, column=col, value=event.event_time or "")
            self._apply_style(cell, cell_style)
            col += 1
            
            # 11. Individuals Involved
            individuals_str = self._format_list(event.participants)
            cell = ws.cell(row=row_idx, column=col, value=individuals_str)
            self._apply_style(cell, cell_style)
            col += 1
            
            # 12. Organizations Involved
            orgs_str = self._format_list(event.organizations)
            cell = ws.cell(row=row_idx, column=col, value=orgs_str)
            self._apply_style(cell, cell_style)
            col += 1
            
            # 13. Casualties (Killed)
            killed = ""
            if event.casualties and "killed" in event.casualties:
                killed = str(event.casualties["killed"])
            cell = ws.cell(row=row_idx, column=col, value=killed)
            self._apply_style(cell, cell_style)
            col += 1
            
            # 14. Casualties (Injured)
            injured = ""
            if event.casualties and "injured" in event.casualties:
                injured = str(event.casualties["injured"])
            cell = ws.cell(row=row_idx, column=col, value=injured)
            self._apply_style(cell, cell_style)
            col += 1
            
            # 15. Source Name
            cell = ws.cell(row=row_idx, column=col, value=event.source_name or "")
            self._apply_style(cell, cell_style)
            col += 1
            
            # 16. Source URL (with hyperlink)
            if event.source_url:
                cell = ws.cell(row=row_idx, column=col, value=event.source_url)
                cell.hyperlink = event.source_url
                cell.font = Font(color=self.LINK_COLOR, underline="single")
                self._apply_style(cell, cell_style)
            else:
                cell = ws.cell(row=row_idx, column=col, value="")
                self._apply_style(cell, cell_style)
            col += 1
            
            # 17. Article Publication Date
            pub_date_str = ""
            if event.article_published_date:
                pub_date_str = event.article_published_date.strftime("%Y-%m-%d")
            cell = ws.cell(row=row_idx, column=col, value=pub_date_str)
            self._apply_style(cell, cell_style)
            col += 1
            
            # 18. Extraction Confidence
            confidence_str = f"{event.confidence:.0%}"
            cell = ws.cell(row=row_idx, column=col, value=confidence_str)
            self._apply_style(cell, cell_style)
            col += 1
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws)
        
        # Set specific column widths for better readability
        ws.column_dimensions['A'].width = 40  # Event Title
        ws.column_dimensions['B'].width = 60  # Summary
        ws.column_dimensions['C'].width = 20  # Event Type
        ws.column_dimensions['D'].width = 25  # Perpetrator
        ws.column_dimensions['E'].width = 35  # Location (Full)
        ws.column_dimensions['F'].width = 20  # City
        ws.column_dimensions['G'].width = 20  # Region/State
        ws.column_dimensions['H'].width = 20  # Country
        ws.column_dimensions['I'].width = 15  # Event Date
        ws.column_dimensions['J'].width = 15  # Event Time
        ws.column_dimensions['K'].width = 30  # Individuals
        ws.column_dimensions['L'].width = 30  # Organizations
        ws.column_dimensions['M'].width = 12  # Casualties (Killed)
        ws.column_dimensions['N'].width = 12  # Casualties (Injured)
        ws.column_dimensions['O'].width = 20  # Source Name
        ws.column_dimensions['P'].width = 50  # Source URL
        ws.column_dimensions['Q'].width = 18  # Publication Date
        ws.column_dimensions['R'].width = 15  # Confidence
        
        # Freeze top row and first column for easier navigation
        ws.freeze_panes = "B2"
        
        logger.info(f"Created Events sheet with {len(events)} rows and {len(headers)} columns")
    
    
    def _create_summary_sheet(self, workbook: Workbook, events: List[EventData]):
        """
        Create a summary/metadata sheet.
        
        Args:
            workbook: Workbook object
            events: List of EventData objects
        """
        ws = workbook.create_sheet("Summary", 1)
        
        # Title
        ws['A1'] = "Event Export Summary"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Export info
        row = 3
        ws[f'A{row}'] = "Export Date:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Total Events:"
        ws[f'B{row}'] = len(events)
        ws[f'B{row}'].font = Font(bold=True)
        
        # Event type breakdown
        row += 2
        ws[f'A{row}'] = "Event Type Breakdown"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        # Count events by type
        type_counts = {}
        for event in events:
            event_type = event.event_type.value
            type_counts[event_type] = type_counts.get(event_type, 0) + 1
        
        row += 1
        ws[f'A{row}'] = "Event Type"
        ws[f'B{row}'] = "Count"
        header_style = self._create_header_style()
        self._apply_style(ws[f'A{row}'], header_style)
        self._apply_style(ws[f'B{row}'], header_style)
        
        for event_type, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
            row += 1
            ws[f'A{row}'] = event_type.upper()
            ws[f'B{row}'] = count
        
        # Location breakdown
        row += 2
        ws[f'A{row}'] = "Top Locations"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        # Count events by location
        location_counts = {}
        for event in events:
            if event.location and event.location.country:
                country = event.location.country
                location_counts[country] = location_counts.get(country, 0) + 1
        
        row += 1
        ws[f'A{row}'] = "Country"
        ws[f'B{row}'] = "Count"
        self._apply_style(ws[f'A{row}'], header_style)
        self._apply_style(ws[f'B{row}'], header_style)
        
        for country, count in sorted(location_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
            row += 1
            ws[f'A{row}'] = country
            ws[f'B{row}'] = count
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws)
        
        logger.info("Created Summary sheet")
    
    def export_to_bytes(
        self,
        events: List[EventData],
        include_metadata: bool = True
    ) -> BytesIO:
        """
        Export events to Excel file in memory (BytesIO).
        
        Args:
            events: List of EventData objects
            include_metadata: Whether to include metadata sheet
        
        Returns:
            BytesIO object containing Excel file
        """
        if not events:
            logger.warning("Attempted to export empty event list")
            raise ValueError("Cannot export empty event list")
        
        # Create workbook
        wb = self.create_events_workbook(events, include_metadata)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Exported {len(events)} events to Excel (BytesIO)")
        return output
    
    def export_to_file(
        self,
        events: List[EventData],
        filepath: str,
        include_metadata: bool = True
    ):
        """
        Export events to Excel file on disk.
        
        Args:
            events: List of EventData objects
            filepath: Path to save Excel file
            include_metadata: Whether to include metadata sheet
        """
        if not events:
            logger.warning("Attempted to export empty event list")
            raise ValueError("Cannot export empty event list")
        
        # Create workbook
        wb = self.create_events_workbook(events, include_metadata)
        
        # Save to file
        wb.save(filepath)
        
        logger.info(f"Exported {len(events)} events to {filepath}")
    
    def get_default_filename(self) -> str:
        """
        Generate a default filename for Excel export.
        
        Returns:
            Filename string with timestamp
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"events_export_{timestamp}.xlsx"
    
    def export_social_events_to_excel(
        self,
        items: List[dict],
        platform_filter: Optional[str] = None
    ) -> BytesIO:
        """
        Export social media search results to Excel, including cached content and analysis.
        
        Args:
            items: List of social event export items with cached data
            platform_filter: Platform name for sheet title (e.g., 'YouTube', 'Twitter')
        
        Returns:
            BytesIO object containing Excel file
        """
        if not items:
            logger.warning("Attempted to export empty social events list")
            raise ValueError("Cannot export empty social events list")
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create social events sheet
        sheet_name = f"{platform_filter} Events" if platform_filter else "Social Media Events"
        ws = wb.create_sheet(sheet_name, 0)
        
        # Define headers for social media export
        headers = [
            "URL",
            "Platform",
            "Title",
            "Snippet",
            "Display Link",
            "Content Cached",
            "Analysis Cached",
            # Cached content fields
            "Posted At",
            "Author Name",
            "Author Username",
            "Verified",
            "Content Text",
            "Likes",
            "Comments",
            "Shares",
            "Views",
            # Cached analysis fields (if available)
            "Event Title",
            "Event Summary",
            "Event Type",
            "Perpetrator",
            "Location (Full Text)",
            "City",
            "Region/State",
            "Country",
            "Event Date",
            "Event Time",
            "Individuals Involved",
            "Organizations Involved",
            "Casualties (Killed)",
            "Casualties (Injured)",
            "Confidence"
        ]
        
        # Write headers
        header_style = self._create_header_style()
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self._apply_style(cell, header_style)
        
        # Write data rows
        row_idx = 2
        for item_data in items:
            is_alt_row = (row_idx % 2 == 0)
            cell_style = self._create_cell_style(is_alt_row)
            
            # Debug logging for first item
            if row_idx == 2:
                logger.info(f"Excel: First item URL: {item_data.get('url', 'N/A')[:50]}...")
                logger.info(f"Excel: cached_content exists: {item_data.get('cached_content') is not None}")
                logger.info(f"Excel: cached_analysis exists: {item_data.get('cached_analysis') is not None}")
                if item_data.get('cached_analysis'):
                    analysis = item_data.get('cached_analysis', {})
                    logger.info(f"Excel: Analysis title: {analysis.get('title', 'N/A')[:80]}...")
                    logger.info(f"Excel: Analysis has location: {analysis.get('location') is not None}")
            
            # Basic social search result data
            ws.cell(row=row_idx, column=1, value=item_data.get('url', ''))
            ws.cell(row=row_idx, column=2, value=item_data.get('platform', ''))
            ws.cell(row=row_idx, column=3, value=item_data.get('title', ''))
            ws.cell(row=row_idx, column=4, value=item_data.get('snippet', ''))
            ws.cell(row=row_idx, column=5, value=item_data.get('display_link', ''))
            
            # Cache status
            cached_content = item_data.get('cached_content')
            cached_analysis = item_data.get('cached_analysis')
            ws.cell(row=row_idx, column=6, value="Yes" if cached_content else "No")
            ws.cell(row=row_idx, column=7, value="Yes" if cached_analysis else "No")
            
            # Cached content details
            if cached_content:
                # Sanitize datetime to remove timezone info (Excel doesn't support timezones)
                posted_at = self._sanitize_datetime_string(cached_content.get('posted_at', ''))
                ws.cell(row=row_idx, column=8, value=posted_at)
                author = cached_content.get('author') or {}  # Handle None author
                ws.cell(row=row_idx, column=9, value=author.get('name', '') if author else '')
                ws.cell(row=row_idx, column=10, value=author.get('username', '') if author else '')
                ws.cell(row=row_idx, column=11, value="Yes" if (author and author.get('verified')) else "No")
                ws.cell(row=row_idx, column=12, value=cached_content.get('text', ''))
                engagement = cached_content.get('engagement') or {}  # Handle None engagement
                ws.cell(row=row_idx, column=13, value=engagement.get('likes', 0) if engagement else 0)
                ws.cell(row=row_idx, column=14, value=engagement.get('comments', 0) if engagement else 0)
                ws.cell(row=row_idx, column=15, value=engagement.get('shares', 0) if engagement else 0)
                ws.cell(row=row_idx, column=16, value=engagement.get('views', 0) if engagement else 0)
            
            # Cached analysis details (extracted event)
            if cached_analysis:
                ws.cell(row=row_idx, column=17, value=cached_analysis.get('title', ''))
                ws.cell(row=row_idx, column=18, value=cached_analysis.get('summary', ''))
                ws.cell(row=row_idx, column=19, value=cached_analysis.get('event_type', ''))
                ws.cell(row=row_idx, column=20, value=cached_analysis.get('perpetrator', ''))
                
                location = cached_analysis.get('location') or {}  # Handle None location
                location_parts = []
                if location and location.get('venue'):
                    location_parts.append(location['venue'])
                if location and location.get('city'):
                    location_parts.append(location['city'])
                if location and location.get('state'):
                    location_parts.append(location['state'])
                if location and location.get('country'):
                    location_parts.append(location['country'])
                location_full = ', '.join(location_parts)
                
                ws.cell(row=row_idx, column=21, value=location_full)
                ws.cell(row=row_idx, column=22, value=location.get('city', '') if location else '')
                ws.cell(row=row_idx, column=23, value=location.get('state', '') if location else '')
                ws.cell(row=row_idx, column=24, value=location.get('country', '') if location else '')
                
                # Sanitize event_date to remove timezone info
                event_date = self._sanitize_datetime_string(cached_analysis.get('event_date', ''))
                ws.cell(row=row_idx, column=25, value=event_date)
                ws.cell(row=row_idx, column=26, value=cached_analysis.get('event_time', ''))
                ws.cell(row=row_idx, column=27, value=self._format_list(cached_analysis.get('participants', [])))
                ws.cell(row=row_idx, column=28, value=self._format_list(cached_analysis.get('organizations', [])))
                
                casualties = cached_analysis.get('casualties') or {}  # Handle None casualties
                ws.cell(row=row_idx, column=29, value=casualties.get('killed', '') if casualties else '')
                ws.cell(row=row_idx, column=30, value=casualties.get('injured', '') if casualties else '')
                ws.cell(row=row_idx, column=31, value=cached_analysis.get('confidence', ''))
            
            # Apply style to all cells in this row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                self._apply_style(cell, cell_style)
            
            row_idx += 1
        
        # Auto-adjust column widths
        self._auto_adjust_column_widths(ws, min_width=12, max_width=60)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Exported {len(items)} social media events to Excel")
        return output


# Global exporter instance
excel_exporter = ExcelExporter()
